import openpyxl

path ="/Users/vaibhavlutade/Downloads"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
col = sheet.max_column
print(rows)
print(col)